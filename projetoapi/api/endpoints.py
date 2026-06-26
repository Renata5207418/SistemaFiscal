#ENDPOINTS EMPRESAS
import re
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Query
from pydantic import BaseModel, constr
from database.database import get_pre_cadastro_collection, get_groups_collection
from typing import Literal, Optional
from login.auth import get_current_user
from utils.utils import limpar_cnpj, validar_cnpj_unico

router = APIRouter()

# --- CONEXÃO COM O BANCO ---
colecao = get_pre_cadastro_collection()
colecao_grupos = get_groups_collection()  

# Modelo atualizado: grupo agora é string livre
class Empresa(BaseModel):
    cod: int
    empresa: str
    cnpj: constr(strip_whitespace=True)
    inscricao_municipal: Optional[str] = ""
    regime_tributario: str
    grupo: str
    ativo: bool = True
    rotina: bool = True


class NovoGrupoRequest(BaseModel):
    tipo: Literal["SN", "LP", "COM"]


# --- FUNÇÃO DE SINCRONIZAÇÃO ---
def sincronizar_grupo(nome_grupo: str):
    """
    Verifica se o grupo existe na coleção de grupos. 
    Se for um grupo novo, insere automaticamente.
    """
    if nome_grupo and nome_grupo.strip() not in ["", "-", "Sem Grupo", "novo_grupo"]:
        nome_limpo = nome_grupo.strip()
        # Busca case-insensitive para não duplicar "1 SN" e "1 sn"
        existe = colecao_grupos.find_one({"grupo": re.compile(f"^{re.escape(nome_limpo)}$", re.IGNORECASE)})
        if not existe:
            colecao_grupos.insert_one({"grupo": nome_limpo})


# --- NOVAS ROTAS DE GRUPOS ---

@router.get("/grupos")
async def listar_grupos(current_user: dict = Depends(get_current_user)):
    docs = list(colecao_grupos.find({}, {"_id": 0, "grupo": 1}))
    lista = [d.get("grupo") for d in docs if d.get("grupo")]

    def sort_key(s):
        s = str(s).strip()
        if s.lower() == "sem grupo": return (999, 999)
        match = re.match(r"(\d+)\s+(.*)", s)
        if match:
            tipo_str = match.group(2).upper()
            pri = 1 if "SN" in tipo_str else 2 if "LP" in tipo_str else 3
            return (pri, int(match.group(1)))
        return (99, s)

    lista.sort(key=sort_key)
    return {"grupos": lista}


@router.post("/grupos/criar")
async def criar_proximo_grupo(
        dados: NovoGrupoRequest,
        current_user: dict = Depends(get_current_user)
):
    tipo = dados.tipo
    regex = f"{tipo}$"
    cursor = colecao_grupos.find({"grupo": {"$regex": regex, "$options": "i"}})

    max_num = 0
    encontrou_algum = False

    for doc in cursor:
        nome_grupo = doc.get("grupo", "")
        match = re.match(r"^(\d+)", nome_grupo)
        if match:
            encontrou_algum = True
            num = int(match.group(1))
            if num > max_num:
                max_num = num

    if not encontrou_algum and max_num == 0:
        novo_numero = 1
    else:
        novo_numero = max_num + 1

    novo_nome = f"{novo_numero} {tipo}"

    if colecao_grupos.find_one({"grupo": novo_nome}):
        raise HTTPException(400, f"Grupo {novo_nome} já existe.")

    colecao_grupos.insert_one({"grupo": novo_nome})
    return {"mensagem": f"Grupo '{novo_nome}' criado com sucesso!", "novo_grupo": novo_nome}


# --- ROTAS DE CADASTRO ---

@router.post("/cadastro")
async def cadastrar_empresa(
        dados: Empresa,
        current_user: dict = Depends(get_current_user)
):
    cnpj_limpo = limpar_cnpj(dados.cnpj)

    if colecao.find_one({"_id": dados.cod}):
        raise HTTPException(status_code=400, detail=f"Já existe um cliente cadastrado com o código {dados.cod}.")

    if colecao.find_one({"cnpj": cnpj_limpo}):
        raise HTTPException(status_code=400, detail=f"O CNPJ {cnpj_limpo} já está vinculado a outro cliente.")

    # Sincroniza o grupo novo (se houver)
    sincronizar_grupo(dados.grupo)

    empresa_doc = {
        "_id": dados.cod,
        "empresa": dados.empresa,
        "cnpj": cnpj_limpo,
        "inscricao_municipal": dados.inscricao_municipal or "",
        "regime_tributario": dados.regime_tributario,
        "grupo": dados.grupo.strip() if dados.grupo else "",
        "ativo": dados.ativo,
        "rotina": dados.rotina,
    }
    
    try:
        colecao.insert_one(empresa_doc)
        return {"mensagem": "Empresa cadastrada com sucesso"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.put("/cadastro/{cod}")
async def atualizar_empresa(
        cod: int,
        dados: Empresa,
        current_user: dict = Depends(get_current_user)
):
    cnpj_limpo = limpar_cnpj(dados.cnpj)
    
    cliente_existente = colecao.find_one({"cnpj": cnpj_limpo})
    if cliente_existente and cliente_existente["_id"] != cod:
        raise HTTPException(status_code=400, detail=f"O CNPJ {cnpj_limpo} já pertence ao cliente Cód {cliente_existente['_id']}.")

    # Sincroniza o grupo novo (se houver)
    sincronizar_grupo(dados.grupo)

    try:
        update_data = {
            "empresa": dados.empresa,
            "cnpj": cnpj_limpo,
            "regime_tributario": dados.regime_tributario,
            "grupo": dados.grupo.strip() if dados.grupo else "",
            "ativo": dados.ativo,
            "rotina": dados.rotina,
        }
        if dados.inscricao_municipal is not None:
            update_data["inscricao_municipal"] = dados.inscricao_municipal

        result = colecao.update_one({"_id": cod}, {"$set": update_data})
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Cliente não encontrado.")
        return {"mensagem": "Empresa atualizada com sucesso"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    

@router.get("/clientes")
async def listar_empresas(
        ativo: bool = Query(True),
        current_user: dict = Depends(get_current_user)
):
    empresas = []
    cursor = colecao.find({"ativo": ativo}).sort("_id", 1)
    for doc in cursor:
        empresas.append({
            "grupo": doc.get("grupo", "Sem grupo"),
            "cod": doc.get("_id"),
            "empresa": doc.get("empresa"),
            "cnpj": doc.get("cnpj"),
            "inscricao_municipal": doc.get("inscricao_municipal", ""),
            "regime_tributario": doc.get("regime_tributario", ""),
            "ativo": doc.get("ativo", True),
            "rotina": doc.get("rotina", True),
        })
    return {"empresas": empresas}


@router.post("/cadastro/upload")
async def upload_empresas(
        file: UploadFile = File(...),
        current_user: dict = Depends(get_current_user)
):
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Envie um arquivo .xlsx ou .xls")
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename=file.file, read_only=True, data_only=True)
        ws = wb.active
        linhas = list(ws.iter_rows(values_only=True))
        header = [str(h).strip().lower() for h in linhas[0]]

        required = ["cod", "empresa", "cnpj", "regime_tributario", "grupo"]
        if any(col not in header for col in required):
            raise HTTPException(400, f"Planilha deve ter colunas: {', '.join(required)}")

        inserted = 0
        erros = []

        for idx, row in enumerate(linhas[1:], start=2):
            data = dict(zip(header, row))
            try:
                cnpj_limpo = limpar_cnpj(str(data["cnpj"] or ""))
                inscricao_val = str(data.get("inscricao_municipal", "") or "").strip()
                grupo_planilha = str(data["grupo"] or "").strip()

                try:
                    validar_cnpj_unico(cnpj_limpo, colecao)
                except:
                    erros.append(f"Linha {idx}: CNPJ duplicado.")
                    continue
                
                # Sincroniza grupos vindos da planilha em lote
                sincronizar_grupo(grupo_planilha)

                doc = {
                    "_id": int(data["cod"]),
                    "empresa": str(data["empresa"] or "").strip(),
                    "cnpj": cnpj_limpo,
                    "inscricao_municipal": inscricao_val,
                    "regime_tributario": str(data["regime_tributario"] or "").strip(),
                    "grupo": grupo_planilha,
                    "ativo": True,
                    "rotina": True,
                }
                colecao.insert_one(doc)
                inserted += 1

            except Exception as e:
                erros.append(f"Linha {idx}: {e}")

        return {"inseridas": inserted, "erros": erros}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Falha ao processar planilha: {e}")
    