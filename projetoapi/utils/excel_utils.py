from io import BytesIO
from datetime import datetime

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils.utils import limpar_cnpj


def formatar_cnpj_excel(cnpj: str) -> str:
    cnpj_limpo = limpar_cnpj(cnpj or "")

    if len(cnpj_limpo) == 14:
        return (
            f"{cnpj_limpo[:2]}."
            f"{cnpj_limpo[2:5]}."
            f"{cnpj_limpo[5:8]}/"
            f"{cnpj_limpo[8:12]}-"
            f"{cnpj_limpo[12:]}"
        )

    return str(cnpj or "")


def formatar_data_excel(valor):
    if not valor:
        return ""

    if isinstance(valor, datetime):
        return valor.strftime("%d/%m/%Y %H:%M")

    if isinstance(valor, str):
        try:
            dt = datetime.fromisoformat(valor.replace("Z", "+00:00"))
            return dt.strftime("%d/%m/%Y %H:%M")
        except Exception:
            return valor

    return str(valor)


def numero_excel(valor, padrao=0.0):
    try:
        if valor is None or valor == "":
            return padrao
        return float(valor)
    except Exception:
        return padrao


def inteiro_excel(valor, padrao=0):
    try:
        if valor is None or valor == "":
            return padrao
        return int(valor)
    except Exception:
        return padrao


def get_nota(nota: dict, *campos, padrao=""):
    if not isinstance(nota, dict):
        return padrao

    for campo in campos:
        valor = nota.get(campo)
        if valor not in [None, ""]:
            return valor

    return padrao


def criar_excel(
    nome_aba: str,
    colunas: list[str],
    linhas: list[list],
    colunas_moeda: list[int] | None = None
):
    colunas_moeda = colunas_moeda or []

    wb = Workbook()
    ws = wb.active
    ws.title = nome_aba[:31]

    ws.append(colunas)

    header_fill = PatternFill("solid", fgColor="D9D9D9")
    header_font = Font(bold=True)
    border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for linha in linhas:
        ws.append(linha)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

            if cell.column in colunas_moeda:
                cell.number_format = 'R$ #,##0.00'

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx, coluna in enumerate(ws.columns, start=1):
        max_len = 0

        for cell in coluna:
            valor = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(valor))

        largura = min(max(max_len + 2, 12), 45)
        ws.column_dimensions[get_column_letter(col_idx)].width = largura

    return wb


def responder_excel(wb: Workbook, nome_arquivo: str):
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{nome_arquivo}"'
        }
    )